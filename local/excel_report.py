"""Excel workbook of closed working-fund periods.

Each time a period is closed, the accounting snapshot is appended as one row to
working-fund-closings.xlsx (created on first use). The workbook is the durable
spreadsheet record the office keeps; the printed one-page report is rendered
separately from the same data.

openpyxl is optional: if it is not installed the caller still saves the JSON
history and prints the report, and append_closing simply returns None.
"""
import os

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except Exception:
    Workbook = None
    load_workbook = None

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE, "working-fund-closings.xlsx")
SHEET = "Closings"

BASE_HEADERS = [
    "Closed At", "Mission", "Period", "Starting Amount", "Expenses",
    "Expected Cash", "Counted Cash", "Discrepancy", "Recorded Txns",
]


def available():
    return Workbook is not None


def denom_label(value, typ):
    return f"{value:,} {typ}"


def denom_key(value, typ):
    """Stable key for a denomination, e.g. (10000, "note") -> "n10000"."""
    return f"{typ[0]}{value}"


def _headers(denoms):
    return BASE_HEADERS + [denom_label(v, t) for v, t in denoms]


def _style_header(ws, headers):
    ws.append(headers)
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2D3C45")
    center = Alignment(horizontal="center")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.fill = fill
        cell.alignment = center
    ws.freeze_panes = "A2"


def append_closing(record, denoms):
    """Append one period-close record. Returns the workbook path, or None if
    openpyxl is unavailable."""
    if Workbook is None:
        return None
    headers = _headers(denoms)
    if os.path.exists(XLSX_PATH):
        wb = load_workbook(XLSX_PATH)
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb.create_sheet(SHEET)
        if ws.max_row <= 1 and ws.cell(row=1, column=1).value is None:
            _style_header(ws, headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET
        _style_header(ws, headers)

    counts = record.get("counts", {})
    row = [
        record.get("closedAt", ""),
        record.get("mission", ""),
        "WF " + str(record.get("period", "")),
        int(record.get("start", 0) or 0),
        int(record.get("spent", 0) or 0),
        int(record.get("expected", 0) or 0),
        int(record.get("counted", 0) or 0),
        int(record.get("discrepancy", 0) or 0),
        int(record.get("recordedCount", 0) or 0),
    ] + [int(counts.get(denom_key(v, t), 0) or 0) for v, t in denoms]
    ws.append(row)
    wb.save(XLSX_PATH)
    return XLSX_PATH
