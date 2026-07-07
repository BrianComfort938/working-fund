import os
import shutil
from datetime import datetime

try:
    import openpyxl
except Exception:
    openpyxl = None

BASE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(BASE)
OUTPUT_DIR = os.path.join(BASE, "closures")

_TEMPLATE_CANDIDATES = [
    os.path.join(BASE, "PRF_Template.xlsx"),
    os.path.join(BASE, "prf_template.xlsx"),
    os.path.join(_ROOT, "PRF_Template.xlsx"),
    os.path.join(_ROOT, "prf_template.xlsx"),
]

def template_path():
    for p in _TEMPLATE_CANDIDATES:
        if os.path.exists(p):
            return p
    return _TEMPLATE_CANDIDATES[0]

PR_SHEET, WF_SHEET, CC_SHEET = "Payment Request", "Working Fund", "Cash Count"
GL_FIRST_ROW, GL_LAST_ROW = 39, 63
COL_B, COL_C, COL_D, COL_E, COL_L, COL_S = 2, 3, 4, 5, 12, 19

CASH_ROWS = [
    (3, 10000, "note"), (4, 5000, "note"), (5, 2000, "note"), (6, 1000, "note"),
    (7, 500, "note"), (8, 500, "coin"), (9, 200, "coin"), (10, 100, "coin"),
    (11, 50, "coin"), (12, 25, "coin"), (13, 10, "coin"), (14, 5, "coin"),
]
WAVE_ROW, ORANGE_ROW = 16, 17

def available():
    return openpyxl is not None and os.path.exists(template_path())

def unavailable_reason():
    if openpyxl is None:
        return "openpyxl is not installed. Run: pip install -r local/requirements.txt"
    if not os.path.exists(template_path()):
        return ("PRF_Template.xlsx was not found. Put the template in the local/ folder "
                "(local/PRF_Template.xlsx) or in the project root.")
    return ""

def cash_denominations():
    return [{"row": r, "value": v, "kind": k} for (r, v, k) in CASH_ROWS]

def _series_acct(account_name):
    head = (account_name or "").split(" ", 1)[0]
    if "-" in head:
        series, acct = head.split("-", 1)
        return series.strip(), acct.strip()
    return None, None

def _label(account_name):
    parts = (account_name or "").split(" ", 1)
    return (parts[1] if len(parts) > 1 else account_name or "").upper()

def fill_close(mission, mission_label, period, account_totals, account_codes,
               cash_counts, wave, orange, date_range=None):
    if not available():
        raise RuntimeError("PRF template or openpyxl unavailable")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    out_name = f"PRF_{mission}_WF{period}_{stamp}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, out_name)
    shutil.copyfile(template_path(), out_path)

    wb = openpyxl.load_workbook(out_path)
    pr, cc = wb[PR_SHEET], wb[CC_SHEET]

    row_by_key, spare_rows = {}, []
    last_line = 0
    for r in range(GL_FIRST_ROW, GL_LAST_ROW + 1):
        d = pr.cell(row=r, column=COL_D).value
        e = pr.cell(row=r, column=COL_E).value
        if d and e:
            row_by_key[(str(d)[-3:], str(e).strip())] = r
            line = pr.cell(row=r, column=COL_B).value
            try:
                last_line = max(last_line, int(line))
            except (TypeError, ValueError):
                pass
        else:
            spare_rows.append(r)
        pr.cell(row=r, column=COL_S).value = None

    written, unmapped, next_spare = [], [], 0
    for code, total in account_totals.items():
        if not total:
            continue
        name = account_codes.get(code, "")
        series, acct = _series_acct(name)
        row = row_by_key.get((series, acct))
        if row is None:
            if next_spare >= len(spare_rows):
                unmapped.append({"code": code, "name": name, "total": int(round(total))})
                continue
            row = spare_rows[next_spare]
            next_spare += 1
            last_line += 1
            pr.cell(row=row, column=COL_B).value = last_line
            pr.cell(row=row, column=COL_C).value = "IVC01"
            pr.cell(row=row, column=COL_D).value = "1385" + (series or "")
            pr.cell(row=row, column=COL_E).value = acct
            pr.cell(row=row, column=COL_L).value = _label(name)
            row_by_key[(series, acct)] = row
        pr.cell(row=row, column=COL_S).value = int(round(total))
        written.append({"code": code, "name": name, "row": row, "total": int(round(total))})

    cash_total = 0
    for (r, value, _kind) in CASH_ROWS:
        qty = cash_counts.get(str(r), cash_counts.get(r, 0)) or 0
        try:
            qty = int(qty)
        except (TypeError, ValueError):
            qty = 0
        pr_qty = max(0, qty)
        cc.cell(row=r, column=COL_B).value = pr_qty
        cash_total += pr_qty * value
    wave, orange = int(wave or 0), int(orange or 0)
    cc.cell(row=WAVE_ROW, column=COL_C).value = wave
    cc.cell(row=ORANGE_ROW, column=COL_C).value = orange

    wf = wb[WF_SHEET]
    if mission_label:
        wf["A3"] = mission_label
    if date_range:
        wf["D3"] = date_range

    wb.save(out_path)
    return {
        "path": out_path, "name": out_name,
        "written": written, "unmapped": unmapped,
        "spentTotal": sum(w["total"] for w in written),
        "countedCash": cash_total, "wave": wave, "orange": orange,
        "countedTotal": cash_total + wave + orange,
    }
